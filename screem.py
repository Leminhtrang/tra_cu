from tkinter import *

from tkinter import ttk

from tkinter import filedialog

import os

import numpy as np

import openpyxl

import xlwings as xw


from tkinter import messagebox


class Setup:

	def __init__(self,root):

		self.root = root

		self.label_frame()

		self.label()

		self.button()

		self.combobox()

		self.treeview()

		style= ttk.Style()
		style.theme_use('clam')

		#style.configure("Treeview.Heading", background="red",relief="sunken",foreground="green")

		

	def label_frame(self):

		self.label_frame_main = LabelFrame(self.root, text="Khu Vực Up Thông Tin Data",relief="sunken",bg="#999933")

		self.label_frame_main.grid(row=1,column=0,sticky="nwes",pady=5,padx=10)

		self.label_frame_main.grid_propagate(False)


		self.label_frame_command = LabelFrame(self.root,relief="ridge",height=70,bg="#90EE90",width=300)

		self.label_frame_command.grid(row=2,column=0,pady=5,padx=5)

		self.label_frame_command.grid_propagate(False)

		self.label_frame_search = LabelFrame(self.label_frame_main,relief="raised",height=80,bg="#9999FF")

		self.label_frame_search.grid(row=0,column=0,sticky="wen",pady=5,padx=5)

		self.label_frame_search.grid_propagate(False)

		self.canvas = Canvas(self.label_frame_search, height=1, background='gray')
		
		self.canvas.grid(row=1, column=0,columnspan=7)

	def label(self):

		self.label_title = Label(self.root,text="Tra Cứu Thông Tin Tỉnh Thành",font = "Times 23 bold",bg="#008080")

		self.label_title.grid(row=0,column=0,sticky="nwe",pady=5)

		self.label_title_treeview = Label(self.label_frame_search,text="Bảng Data Thông Tin",font = "Times 11 bold",bg="#C71585")

		self.label_title_treeview.grid(row=2,column=0,sticky="wes",pady=5,padx=5,columnspan=7)




		self.label_list_title = Label(self.label_frame_search,text="Tiêu Đề ",bg="#EEAEEE")

		self.label_list_title.grid(row=0,column=2,pady=5,padx=5)

		self.label_list_value_title = Label(self.label_frame_search,text="Giá Trị Tiêu Đề ",bg="#EEAEEE")

		self.label_list_value_title.grid(row=0,column=4,pady=5,padx=5)



	def button(self):


		self.button_log_file = Button(self.label_frame_search,text="Log In File",font="Times 9 bold",borderwidth=2,bg="#FFFF00")

		self.button_log_file.grid(row=0,column=0,pady=5,padx=5)


		self.button_enter = Button(self.label_frame_command,text="Xuất File Excel",font="Times 11 bold",borderwidth=5,bg="#3366CC",width=12,height=2)

		self.button_enter.grid(row=0,column=1,sticky="ns",pady=5)

		self.button_delete = Button(self.label_frame_command,text="Xóa",font="Times 11 bold",borderwidth=2,relief="groove",bg="#FF66FF",width=7)

		self.button_delete.grid(row=0,column=0,sticky="ws",padx=5,pady=5)


		self.button_quit = Button(self.label_frame_command,text="Quit",font="Times 11 bold",borderwidth=2,relief="groove",bg="red",width=7)

		self.button_quit.grid(row=0,column=2,sticky="es",padx=5,pady=5)


	def combobox(self):

		self.combobox_sheet_excel = ttk.Combobox(self.label_frame_search,values=["Sheet Excel(Nếu Có)"],state="disable",justify="center")

		self.combobox_sheet_excel.current(0)

		self.combobox_sheet_excel.grid(row=0, column=1, padx=5, pady=5)


		self.combobox_list_title = ttk.Combobox(self.label_frame_search,state="readonly",justify="center")

		self.combobox_list_title.grid(row=0, column=3, padx=5, pady=5)

		self.combobox_list_value_title = ttk.Combobox(self.label_frame_search,state="disable",justify="center")

		self.combobox_list_value_title.grid(row=0, column=5, padx=5, pady=5)

		self.combobox_list_value_title["values"] = "???"

		self.combobox_list_value_title.current(0)


	def treeview(self):

		self.frame = Frame(self.label_frame_main)

		self.frame.grid(row=1,column=0,sticky="nsew")

		self.treeview =ttk.Treeview(self.frame,show="headings",selectmode="none")

		self.treeview.grid(row=0,column=0,sticky="nsew")


		self.xscrollbar = ttk.Scrollbar(self.frame, orient="horizontal", command=self.treeview.xview)

		self.treeview.configure(xscrollcommand=self.xscrollbar.set)

		self.xscrollbar.grid(row=1,column=0,sticky="wes")


		self.yscrollbar = ttk.Scrollbar(self.frame, orient="vertical", command=self.treeview.yview)

		self.treeview.configure(yscrollcommand=self.yscrollbar.set)

		self.yscrollbar.grid(row=0,column=1,sticky="sen")

		def select(event=None):
			self.treeview.selection_toggle(self.treeview.focus())

		
		self.treeview.bind("<ButtonRelease-1>", select)
		



class Balance(Setup):

	def __init__(self, root):

		super().__init__(root)

		self.balance_widget()


	def balance_widget(self):


		def function_balance(main,lis,wei_,min_):

			for widget in lis:

				Grid.rowconfigure(main, widget, weight=wei_, minsize=min_)
				Grid.columnconfigure(main, widget, weight=wei_, minsize=min_)

		def function_balance2(main,widget):

			Grid.rowconfigure(main, widget, weight=1, minsize=0)
			Grid.columnconfigure(main, widget, weight=1, minsize=0)

		lis_widget_search = [self.label_list_title,self.label_list_value_title,self.combobox_list_title,self.combobox_list_value_title,

		self.button_log_file,self.combobox_sheet_excel,self.label_title_treeview]


		lis_widget_command = [self.button_enter,self.button_delete,self.button_quit]

		
		function_balance(self.label_frame_search,lis_widget_search,1,0)

		function_balance(self.label_frame_command,lis_widget_command,1,0)

		function_balance2(self.root,self.label_frame_main)

		function_balance2(self.label_frame_main,self.frame)


		function_balance2(self.frame,self.treeview)


		

class funcition(Balance):

	def __init__(self, root):

		super().__init__(root)

		self.button_log_file.config(command=self.upfile_data)	

		self.button_enter.config(command=self.enter)

		self.button_delete.config(command=self.delete)

		self.button_quit.config(command=self.quit)

		self.combobox_list_title.bind("<<ComboboxSelected>>",self.take_value_title)

		self.combobox_list_value_title.bind("<<ComboboxSelected>>",self.filter_data)

		self.combobox_sheet_excel.bind("<<ComboboxSelected>>",self.change_data)

		self.treeview.bind("<Control-a>", self.select_all)

		self.root.bind("<Double-Button-1>", self.clear_selection)




	def upfile_data(self):

		self.filename = filedialog.askopenfilename(initialdir=os.getcwd(),title="Chọn File Excel",filetypes=[("Excel files", ".xlsx .xlsm .xltx .xltm .xlt .csv")])
	
		if self.filename =="":

			return

		else:

			for item in self.treeview.get_children():

				self.treeview.delete(item)

			if os.path.splitext(os.path.basename(self.filename))[1] == ".csv":

				self.data = np.loadtxt(self.filename, delimiter=',', encoding='utf-8', dtype=str)

				self.data = [[item.replace('"', '') for item in sublist] for sublist in self.data.tolist()]


			else:

				self.combobox_sheet_excel["state"] = "readonly"

				self.data_raw,lis_sheet= self.upfile_data_excel(self.filename,self.combobox_sheet_excel)

				self.combobox_sheet_excel["values"] = lis_sheet

				self.combobox_sheet_excel.current(0)



				self.data_handle = self.data_raw[self.combobox_sheet_excel.current()]

				for i in range(self.data_handle.shape[0]):
					for j in range(self.data_handle.shape[1]):
						if self.data_handle[i, j] is None:
							self.data_handle[i, j] = ""


				self.data = self.data_handle.tolist()

			if len(self.data) <=1:

			

				self.treeview["columns"] = ("Name")

				self.treeview.heading("#1", text="Phần Nầy không Có Data Hoặc Bạn Đã Xóa Hết Data",anchor="center")

				self.treeview.column("Name", width=self.treeview.winfo_width())

				self.button_enter["state"] = "disable"

				self.button_delete["state"] = "disable"

				
				self.treeview.unbind("<Double-Button-1>")

			else:
				
				self.treeview.bind("<Double-Button-1>", self.edit_treeview)

				self.button_enter["state"] = "normal"

				self.button_delete["state"] = "normal"


				self.treeview["columns"] = tuple(self.data[0])

				for col_name in self.treeview["columns"]:

					self.treeview.heading(col_name, text=col_name)


				for values in self.data[1:]:

					
					self.treeview.insert("", "end", values=values)	

				self.minsize = self.treeview.winfo_width()//len(self.treeview['columns'])

				for i in self.treeview['columns']:

					self.treeview.column(i, stretch=False,minwidth =self.minsize)	


				self.color_even_rows(self.treeview)


				self.combobox_list_title["values"]= ["Tất Cả"] +self.data[0]

				self.combobox_list_title.current(0)

				

				matrix = np.array(self.data, dtype=object)[1:]

				self.row_fill = np.arange(len(matrix))

				self.treeview.bind("<Double-Button-1>", self.edit_treeview)



	def change_data(self,event):

		for item in self.treeview.get_children():

			self.treeview.delete(item)

		self.data = self.data_raw[self.combobox_sheet_excel.current()].tolist()

		if len(self.data) <=1:

			

			self.treeview["columns"] = ("Name")

			self.treeview.heading("#1", text="Phần Nầy không Có Data Hoặc Bạn Đã Xóa Hết Data",anchor="center")

			self.treeview.column("Name", width=self.treeview.winfo_width())

			self.button_enter["state"] = "disable"

			self.button_delete["state"] = "disable"

			
			self.treeview.unbind("<Double-Button-1>")	

		else:


			self.treeview.bind("<Double-Button-1>", self.edit_treeview)

			self.button_enter["state"] = "normal"

			self.button_delete["state"] = "normal"

			for item in self.treeview.get_children():

				self.treeview.delete(item)

			self.treeview["columns"] = tuple(self.data[0])

			for col_name in self.treeview["columns"]:

				self.treeview.heading(col_name, text=col_name)


			for values in self.data[1:]:

				
				self.treeview.insert("", "end", values=values)

			self.minsize = self.treeview.winfo_width()//len(self.treeview['columns'])

			for i in self.treeview['columns']:

				self.treeview.column(i, stretch=False,minwidth =self.minsize)	


			self.color_even_rows(self.treeview)


			self.combobox_list_title["values"]= ["Tất Cả"] +self.data[0]

			self.combobox_list_title.current(0)



			self.combobox_list_value_title["values"] = "???"

			self.combobox_list_value_title.current(0)

			self.combobox_list_value_title["state"] = "disable"

			matrix = np.array(self.data, dtype=object)[1:]

			self.row_fill = np.arange(len(matrix))



	def take_value_title(self,event):

		matrix = np.array(self.data, dtype=str)[1:]



		first_column = matrix[:, self.combobox_list_title.current()-1]

		
			
		
		unique_values, indices = np.unique(first_column, return_index=True)

		# Sắp xếp theo thứ tự xuất hiện

		sorted_indices = np.argsort(indices)



		
		unique_values_sorted = unique_values[sorted_indices].tolist()
		

		if self.combobox_list_title.get() == "Tất Cả":

			self.row_fill = np.arange(len(matrix))

			if self.combobox_list_value_title.get() =="???":

				return

			else:

				self.combobox_list_value_title["values"] = "???"

				self.combobox_list_value_title.current(0)

				self.combobox_list_value_title["state"] = "disable"

				for item in self.treeview.get_children():

					self.treeview.delete(item)

				for values in matrix.tolist():
							
					self.treeview.insert("", "end", values=values)

				self.color_even_rows(self.treeview)

		else:

			self.combobox_list_value_title["values"] = unique_values_sorted

			self.combobox_list_value_title.current(0)

			self.combobox_list_value_title["state"] = "readonly"

	def filter_data(self,event):

		if self.combobox_list_value_title.get() == "???":

			return

		else:

			matrix = np.array(self.data, dtype=str)[1:]

			self.row_fill = np.where(matrix[:,self.combobox_list_title.current()-1] == self.combobox_list_value_title.get())[0]



			for item in self.treeview.get_children():

				self.treeview.delete(item)

			data_fill = matrix[self.row_fill, :].tolist()

			for values in data_fill:
						
				self.treeview.insert("", "end", values=values)

			self.color_even_rows(self.treeview)




	def edit_treeview(self,event):

		

		if self.treeview.identify_region(event.x, event.y) == 'cell':

			
			def ok(event):
		


				self.treeview.set(item, column, entry.get())

				value = entry.get()

				entry.destroy()

				column_change = self.row_fill[int(self.treeview.index(item))]

				row_change = int(column[1:])-1

				chang_matrix = np.array(self.data)
				
				
				chang_matrix[column_change+1,row_change] = value
 
				self.data = chang_matrix.tolist()

			column = self.treeview.identify_column(event.x)  # identify column

			item = self.treeview.identify_row(event.y)  # identify item
			x, y, width, height = self.treeview.bbox(item, column) 
			value = self.treeview.set(item, column)

			



		elif self.treeview.identify_region(event.x, event.y) == 'heading': 
			# the user clicked on a heading

			def ok(event):
				
				self.treeview.heading(column, text=entry.get())

				value = entry.get()
				entry.destroy()

				
				
				column_change = 0

				row_change = int(column[1:])-1

				chang_matrix = np.array(self.data)
				
				
				chang_matrix[column_change,row_change] = value
 
				self.data = chang_matrix.tolist()

			column = self.treeview.identify_column(event.x) # identify column
			# tree.bbox work sonly with items so we have to get the bbox of the heading differently
			x, y, width, _unused_ = self.treeview.bbox(self.treeview.get_children('')[0], column) # get x and width (same as the one of any cell in the column)
			# get vertical coordinates (y1, y2)
			y2 = y
			# get bottom coordinate
			while self.treeview.identify_region(event.x, y2) != 'heading':  
				y2 -= 1
			# get top coordinate
			y1 = y2
			while self.treeview.identify_region(event.x, y1) == 'heading':
				y1 -= 1
			height = y2 - y1
			y = y1
			value = self.treeview.heading(column, 'text')

		elif self.treeview.identify_region(event.x, event.y) == 'nothing': 


			column = self.treeview.identify_column(event.x) # identify column
			# check whether we are below the last row:
			x, y, width, height = self.treeview.bbox(self.treeview.get_children('')[-1], column)
			if event.y > y:

				def ok(event):
		
					self.treeview.set(item, column, entry.get())

					value = entry.get()

					entry.destroy()

					column_change = int(self.treeview.index(item))+1

					row_change = int(column[1:])-1

					chang_matrix = np.array(self.data)
					
					
					chang_matrix[column_change,row_change] = value
	 
					self.data = chang_matrix.tolist()

				y += height
				value = ""
			else:
				return
		else:
			return

		# display the Entry   

		entry = Entry(self.treeview,bg="#FFCC00")  # create edition entry
		
		entry.place(x=x, y=y, width=width, height=height,
					anchor='nw')  # display entry on top of cell
		entry.insert(0, value)  # put former value in entry





		entry.bind('<FocusOut>', ok)  # validate with Enter
		entry.focus_set()

	
	def color_even_rows(self,tree):
		for i, item in enumerate(tree.get_children()):
			if i % 2 == 0:
				tree.item(item, tags=("even_row",))
		tree.tag_configure("even_row", background="#CDC9A5")
		


	def clear_selection(self, event):

		self.treeview.selection_remove(self.treeview.selection())

	def select_all(self,event):

		for item in self.treeview.get_children():
			self.treeview.selection_add(item)

	
		

	def calculate_total_column_width(self,treeview):
		total_width = 0
		for column in treeview["columns"]:
			column_width = treeview.column(column, "width")
			total_width += column_width
		return total_width


	def upfile_data_excel(self,filename,combobox):
	

		wb = openpyxl.load_workbook(filename, data_only=True)

		all_data = []
		sheet_names = []

		for sheet in wb.sheetnames:
			current_sheet = wb[sheet]
			num_rows = current_sheet.max_row
			num_cols = current_sheet.max_column
			sheet_data = []

			for i, row in enumerate(current_sheet.iter_rows(values_only=True)):
				# Kiểm tra xem hàng có chứa dữ liệu không
				if any(row):
					sheet_data.append([value for value in row])

			sheet_data = np.array(sheet_data, dtype=object)
			column_has_data = np.any(sheet_data != None, axis=0)
			sheet_data = sheet_data[:, column_has_data]

			# Thêm dữ liệu và tên sheet vào danh sách (loại bỏ các hàng và cột trống thừa)
			all_data.append(sheet_data)
			sheet_names.append(sheet)			

		combobox["values"] = sheet_names

		combobox.current(0)


		return (all_data,sheet_names)


	def enter(self):
		try:
			result = messagebox.askyesno("Confirmation", "Do you want to proceed?")

			if result:

				value = self.data

				wb =  xw.Book()

				sht = wb.sheets[0]

				sht.range("A1").value = value

				sht.api.UsedRange.EntireColumn.AutoFit()

				for edge in [7,8,9,10,11,12]:

					sht.range(sht.used_range).api.Borders(edge).LineStyle = 1


			else:

				return
		except AttributeError:

			pass
	def quit(self):

		self.root.destroy()

	def delete(self):
	
		try:

			rows_to_delete = []

			selected_item = self.treeview.selection()

			for x in selected_item:

				column_change = self.row_fill[int(self.treeview.index(x))]

				rows_to_delete.append(column_change)

			


			

			row_remaining = np.setxor1d(self.row_fill, rows_to_delete) 

			

			for item in self.treeview.get_children():

					self.treeview.delete(item)


			list_rows_to_delete = np.array(rows_to_delete)+1
			

			self.data = np.delete(self.data, list(reversed(list_rows_to_delete)), axis=0)

			if os.path.splitext(os.path.basename(self.filename))[1] != ".csv":

			
				self.data_raw[self.combobox_sheet_excel.current()] = self.data

			
			matrix = np.array(self.data, dtype=object)[1:]
			

			if self.combobox_list_value_title.get() == "???":

				self.row_fill = np.arange(len(matrix))

				for values in matrix.tolist():

					
					self.treeview.insert("", "end", values=values)
			else:
		

				self.row_fill = np.where(matrix[:,self.combobox_list_title.current()-1] == self.combobox_list_value_title.get())[0]

				data_fill = matrix[self.row_fill, :].tolist()

				for values in data_fill:
							
					self.treeview.insert("", "end", values=values)

			rows_to_delete.clear()

			self.color_even_rows(self.treeview)

		except AttributeError:

			pass

if __name__ == "__main__":

	root = Tk()
	funcition(root)



	root.geometry("700x560")

	root.title("Tra Cú Thông Tin")

	root["bg"] ="#008080"

	root.mainloop()